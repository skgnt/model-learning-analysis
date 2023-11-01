from flask import Flask, render_template, g, request, jsonify,send_file
#corsを使うために必要
from flask_cors import CORS
import sqlite3
import openpyxl
import subprocess
import os
import webbrowser

app = Flask(__name__, template_folder='./frontend/build',
            static_folder='./frontend/build/static')
CORS(app)



db_name = r"log/analyze_record.db"

app.json.ensure_ascii = False


@app.route('/', defaults={'path': ''})
@app.route('/<path:path>')
def home(path):
    return render_template('index.html'), 200

# /api/group_infoにアクセスしたときrun_nameとgroup_nameのjsonを返す
# 返すときはjson形式にする


@app.route("/api/group_info")
def group_info():

    sqlite = Sqlite_ctr(get_db())
    return jsonify(sqlite.get_run_group())

# /api/get_log_data
# typeによって返すデータを変える
@app.route("/api/get_roc_data")
def get_roc_data():
    query_params = request.args
    run_name = query_params.get("run_name")
    group_name = query_params.get("group_name")
    log_path = query_params.get("log")
    #log_pathの_を/に変換する
    log_path = log_path.replace("---","/")
    app.logger.info(log_path)
    roc_path = rf"{log_path}/{run_name}/{run_name}_test_roc.png"
    return send_file(roc_path, mimetype='image/png')

@app.route("/api/get_log_data", methods=["GET", "POST"])
def get_log_data():

    query_params = request.get_json()
    type = query_params.get("type")
    run_name = query_params.get("run_name")
    log_path = query_params.get("log")
    if type == "confusion_matrix":
        # confusion_matrixのcsvデータを返す
        # クエリのlogを取得する
        

        try:
            cm_list = read_excel_file(
                rf"{log_path}/{run_name}/{run_name}_test_cm.xlsx")
        except:
            cm_list = read_excel_file(
                rf"{log_path}/{run_name}/{run_name}_test_bc_cm.xlsx")
        # cm_listをcsv形式にする
        csv_data = ""
        for row in cm_list:
            csv_data += ",".join(str(cell) for cell in row) + "\n"
        # Remove the newline character from the end of the CSV data
        csv_data = csv_data.rstrip('\n')
        return jsonify({"data": csv_data})
    elif type == "individual_data":
        #log_path直下のrun_name_test.csvを返す
        csv_path = rf"{log_path}/{run_name}/{run_name}_test_p.csv"

        return send_file(csv_path, mimetype='text/csv')


#画像のpathがクエリに含まれているので、その画像を返す
@app.route("/api/get_image")
def get_image():
    query_params = request.args
    image_path = query_params.get("path")
    image_path = image_path.replace("---","/")
    return send_file(image_path, mimetype='image/png')



#画像のpathがクエリに含まれているので、その画像を開く
@app.route("/api/open_image")
def open_image():
    query_params = request.args
    image_path = query_params.get("path")
    image_path = image_path.replace("---","/")
    print(image_path)
    os.system(f'"{image_path}"')

    return jsonify({"status": "ok"})


# /api/group_data/group_nameにアクセスしたときgroup_nameを指定してデータを取得する
# group_nameは可変で、string型で指定する
# 返すときはjson形式にする
@app.route("/api/get_statistics_data")
def get_statistics_data():
    # クエリでgroup_name、run_nameを取得する。
    # run_nameがクエリに存在しない場合はget_group_dataを呼び出す
    # run_nameがクエリに存在する場合はget_run_dataを呼び出す
    group_name = request.args.get("group_name")
    run_name = request.args.get("run_name")
    # 現在のpathを取得する
    

    sqlite = Sqlite_ctr(get_db())
    if run_name == None:
        return jsonify(sqlite.get_group_data(group_name))
    else:
        return jsonify(sqlite.get_run_data(run_name, group_name))


def read_excel_file(path):
    # excelファイルを開く
    wb = openpyxl.load_workbook(path)
    # confusion_matrixのシートを開く
    ws = wb['confusion_matrix']
    # b2からセルの端までの範囲を取得して、配列にする
    ws_range = ws['b2':f'{ws.cell(ws.max_row,ws.max_column).coordinate}']

    # 2次元配列にする
    ws_list = []

    for row in ws_range:
        row_list = []
        for cell in row:
            row_list.append(cell.value)
        ws_list.append(row_list)
    ws_list[0][0] = " "
    return ws_list


@app.route("/api/open_explorer")
def open_explorer():
    path = request.args.get("log")
    path = path.replace("---", "\\")
    app.logger.info(path)
    try:
        subprocess.run(["explorer", r"{}".format(path)], timeout=5, check=True)
        return jsonify({"status": "ok"})
    except subprocess.TimeoutExpired:
        return jsonify({"status": "error", "message": "Timeout expired"})
    except subprocess.CalledProcessError:
        #CalledProcessErroの場合は、正常として処理する。
        return jsonify({"status": "ok"})
    except:
        return jsonify({"status": "error", "message": "Unexpected error"})


# sqlite3の基本的なクラス
class Sqlite_ctr:
    def __init__(self, conn):
        self.conn = conn
        self.cur = self.conn.cursor()

    def select(self, sql):
        self.cur.execute(sql)
        return self.cur.fetchall()

    # test_
    # resultテーブルからrun_nameとgroup_nameを取得する

    def get_run_group(self):
        sql = "select run_name, group_name from test_result"
        # {"group_name": {"run_name": "run_name"}}というdict形式にする
        # ただし、group_nameは同じものは同じgroup_nameのkeyの中にrun_nameを追加する
        res_dict = {}
        res = self.select(sql)

        for run_name, group_name in res:
            if group_name not in res_dict:
                res_dict[group_name] = {}
            res_dict[group_name][run_name] = run_name

        # json形式に変換して返す
        return res_dict

    # group_nameが一致するものを取得する
    # ただし、データがないnullは削除して取得する。
    def get_group_data(self, group_name):
        sql = "select * from test_result where group_name = '{}'".format(
            group_name)
        res = self.select(sql)

        # headerを取得する
        header = [column[0] for column in self.cur.description]

        # nullを削除しながら,dict形式にする
        # {index: {column: value}}という形式にする
        res_dict = {}
        for index, ct in enumerate(res):
            res_dict["keyID"+str(index)] = {}
            for i in range(len(ct)):
                if ct[i] != None:
                    # ct[i]がfloat型の場合は少数第5位までにする
                    if type(ct[i]) == float:
                        res_dict["keyID"+str(index)][header[i]
                                                     ] = round(ct[i], 5)
                    else:
                        res_dict["keyID"+str(index)][header[i]] = str(ct[i])

        return res_dict
    # run_nameとgroup_nameが一致するものを取得する
    # ただし、データがないnullは削除して取得する。

    def get_run_data(self, run_name, group_name):
        sql = "select * from test_result where run_name = '{}' and group_name = '{}'".format(
            run_name, group_name)
        res = self.select(sql)

        # headerを取得する
        header = [column[0] for column in self.cur.description]

        # nullを削除しながら,dict形式にする
        # {index: {column: value}}という形式にする
        res_dict = {}
        for index, ct in enumerate(res):
            res_dict["keyID"+str(index)] = {}
            for i in range(len(ct)):
                if ct[i] != None:
                    # ct[i]がfloat型の場合は少数第5位までにする
                    if type(ct[i]) == float:
                        res_dict["keyID"+str(index)][header[i]
                                                     ] = round(ct[i], 5)
                    else:
                        res_dict["keyID"+str(index)][header[i]] = str(ct[i])

        return res_dict


def get_db():
    if 'db' not in g:
        g.db = sqlite3.connect(db_name)
    return g.db


def close_db(exception=None):
    db = g.pop('db', None)

    if db is not None:
        db.close()


if __name__ == "__main__":
    # sqlite = Sqlite_ctr(sqlite3.connect(db_name))
    # print(sqlite.get_group_data("cifar10_3b"))
    webbrowser.open("http://localhost:8000")
    app.run(debug=True, port=8000)

    
